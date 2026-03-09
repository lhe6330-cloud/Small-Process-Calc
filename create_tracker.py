"""
创建 V2.0 任务追踪 Excel 文件（2.5 小时版）
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = '任务追踪'

# 定义样式
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = [
    '任务 ID', '任务名称', '阶段', '预计耗时 (分钟)', 
    '计划开始', '计划结束',
    '实际开始', '实际结束', '实际耗时 (分钟)',
    '状态', '积分'
]

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 任务数据（12 个任务，2.5 小时版）
tasks = [
    # 阶段一：后端开发
    ('T1', 'VLE 模块', '后端开发', 20, '12:15', '12:35'),
    ('T2', '分离器模块', '后端开发', 20, '12:35', '12:55'),
    ('T3', '涡轮一维设计', '后端开发', 20, '12:55', '13:15'),
    # 午休
    ('BREAK', '午休', '休息', 10, '13:15', '13:25'),
    # 阶段二：API 集成
    ('T4', '流程节点+VLE API', 'API 集成', 10, '13:25', '13:35'),
    ('T5', '分离器 + 涡轮 API', 'API 集成', 15, '13:35', '13:50'),
    # 阶段三：前端开发
    ('T6', '分离器位置 UI', '前端开发', 15, '13:50', '14:05'),
    ('T7', '涡轮设计表单', '前端开发', 15, '14:05', '14:20'),
    ('T8', '结果面板联动', '前端开发', 10, '14:20', '14:30'),
    # 阶段四：报告导出
    ('T9', '分离器 PDF/Excel', '报告导出', 10, '14:30', '14:40'),
    ('T10', '涡轮设计 PDF/Excel', '报告导出', 10, '14:35', '14:45'),
    # 阶段五：测试验收
    ('T11', '后端单元测试', '测试调试', 10, '14:30', '14:40'),
    ('T12', '全链路验收', '测试调试', 5, '14:40', '14:45'),
]

# 写入任务数据
for row, task in enumerate(tasks, 2):
    # 基本信息
    ws.cell(row=row, column=1, value=task[0]).border = thin_border
    ws.cell(row=row, column=2, value=task[1]).border = thin_border
    ws.cell(row=row, column=3, value=task[2]).border = thin_border
    ws.cell(row=row, column=4, value=task[3]).border = thin_border
    ws.cell(row=row, column=5, value=task[4]).border = thin_border
    ws.cell(row=row, column=6, value=task[5]).border = thin_border
    
    # 实际时间（留空，待填写）
    ws.cell(row=row, column=7, value='').border = thin_border
    ws.cell(row=row, column=8, value='').border = thin_border
    ws.cell(row=row, column=9, value='').border = thin_border
    
    # 状态（下拉列表）
    status_cell = ws.cell(row=row, column=10)
    status_cell.border = thin_border
    status_cell.value = '未开始'
    
    # 积分
    points_cell = ws.cell(row=row, column=11)
    points_cell.border = thin_border
    points_cell.value = 0
    
    # 午休行特殊样式
    if task[0] == 'BREAK':
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')

# 设置列宽
column_widths = [10, 22, 12, 14, 12, 12, 12, 12, 14, 10, 10]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 添加数据验证（状态下拉列表）
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"未开始，进行中，已完成"', allow_blank=True)
dv.error = '请选择有效的状态'
dv.errorTitle = '无效状态'
ws.add_data_validation(dv)
dv.add('J2:J14')

# 添加积分规则说明表
ws2 = wb.create_sheet(title='积分规则')
rules_data = [
    ['V2.0 开发任务 - 积分规则', ''],
    ['', ''],
    ['积分计算规则：', ''],
    ['快速完成（实际耗时 ≤ 0.5 × 预计耗时）', '+1.5 分'],
    ['按时完成（0.5 × 预计耗时 < 实际耗时 ≤ 预计耗时）', '+1 分'],
    ['超时完成（预计耗时 < 实际耗时 ≤ 2 × 预计耗时）', '+0.5 分'],
    ['严重超时（实际耗时 > 2 × 预计耗时）', '-1 分'],
    ['', ''],
    ['积分公式（K 列）：', ''],
    ['在 K2 单元格输入以下公式，然后向下填充：', ''],
    ['=IF(I2="", 0, IF(I2<=D2*0.5, 1.5, IF(I2<=D2, 1, IF(I2<=D2*2, 0.5, -1))))', ''],
    ['', ''],
    ['总分计算：', '=SUM(K2:K14)'],
]

for row, rule in enumerate(rules_data, 1):
    for col, val in enumerate(rule, 1):
        ws2.cell(row=row, column=col, value=val)

ws2.column_dimensions['A'].width = 50
ws2.column_dimensions['B'].width = 30

# 添加统计摘要表
ws3 = wb.create_sheet(title='统计摘要')
stats_data = [
    ['V2.0 开发任务 - 统计摘要', ''],
    ['', ''],
    ['任务总数：', '=COUNTA(A2:A14)-1'],
    ['已完成：', '=COUNTIF(J2:J14,"已完成")'],
    ['进行中：', '=COUNTIF(J2:J14,"进行中")'],
    ['未开始：', '=COUNTIF(J2:J14,"未开始")'],
    ['', ''],
    ['总积分：', '=SUM(K2:K14)'],
    ['平均积分：', '=AVERAGE(K2:K14)'],
    ['', ''],
    ['阶段统计：', ''],
    ['后端开发积分：', '=SUMIF(C2:C14,"后端开发",K2:K14)'],
    ['API 集成积分：', '=SUMIF(C2:C14,"API 集成",K2:K14)'],
    ['前端开发积分：', '=SUMIF(C2:C14,"前端开发",K2:K14)'],
    ['报告导出积分：', '=SUMIF(C2:C14,"报告导出",K2:K14)'],
    ['测试调试积分：', '=SUMIF(C2:C14,"测试调试",K2:K14)'],
]

for row, stat in enumerate(stats_data, 1):
    for col, val in enumerate(stat, 1):
        ws3.cell(row=row, column=col, value=val)

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 30

# 保存文件
filename = 'C:/Users/Administrator/openclaw-workspace/Small-Process-Calc/V2.0_Task_Tracker.xlsx'
wb.save(filename)
print('OK: Excel file updated:', filename)
print('Task count: 12')
print('Total time: 150 minutes (2.5 hours)')
