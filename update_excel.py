from openpyxl import load_workbook
from datetime import datetime

# 加载 Excel 文件
file_path = r'C:\Users\Administrator\openclaw-workspace\Small-Process-Calc\任务跟踪表.xlsx'
wb = load_workbook(file_path)
ws = wb['任务跟踪表']
ws_score = wb['积分统计']

# 当前时间
now = datetime.now().strftime('%Y-%m-%d %H:%M')

# 更新 T01: 创建项目目录结构
# 行号 = 任务行 + 1 (因为第 1 行是表头)
# T01 在第 2 行，T02 在第 3 行，以此类推
# 格式：[开始时间，完成时间，实际工时，状态，积分]

updates = {
    # T01-T06 Phase 1 已完成
    2: ['00:16', '00:25', 0.15, '✅ 已完成', 1.5],  # T01
    3: ['00:25', '00:32', 0.12, '✅ 已完成', 1.5],  # T02
    4: ['00:32', '00:50', 0.3, '✅ 已完成', 1.5],   # T03
    5: ['00:50', '01:05', 0.25, '✅ 已完成', 1.5],  # T04
    6: ['01:05', '01:15', 0.17, '✅ 已完成', 1.5],  # T05
    7: ['01:15', '01:20', 0.08, '✅ 已完成', 1.5],  # T06 - 刚刚完成
}

for row, data in updates.items():
    ws[f'E{row}'] = data[0]  # 开始时间
    ws[f'F{row}'] = data[1]  # 完成时间
    ws[f'G{row}'] = data[2]  # 实际工时
    ws[f'H{row}'] = data[3]  # 状态
    ws[f'I{row}'] = data[4]  # 积分

# 更新 T07 进行中
ws['E8'] = '01:20'  # 开始时间
ws['H8'] = '🔄 进行中'  # 状态

# 更新积分统计
# 当前总积分 = SUM(I2:I36)
# 已完成数量
completed = sum(1 for row in range(2, 37) if ws[f'H{row}'].value == '✅ 已完成')
in_progress = sum(1 for row in range(2, 37) if ws[f'H{row}'].value == '🔄 进行中')
not_started = 35 - completed - in_progress

ws_score['B12'] = 35  # 总任务数
ws_score['B13'] = completed  # 已完成
ws_score['B14'] = in_progress  # 进行中
ws_score['B15'] = not_started  # 未开始
# B16 是公式 =SUM('任务跟踪表'!I2:I36)
ws_score['B19'] = f'{completed}/35'  # 完成率

# 更新阶段进度
# Phase 1: 6 个任务 (行 2-7)
p1_completed = sum(1 for row in range(2, 8) if ws[f'H{row}'].value == '✅ 已完成')
ws_score['C25'] = p1_completed
ws_score['D25'] = f'{p1_completed/6*100:.0f}%'

# Phase 2: 8 个任务 (行 8-15)
p2_in_progress = sum(1 for row in range(8, 16) if ws[f'H{row}'].value == '🔄 进行中')
ws_score['C26'] = p2_in_progress
ws_score['D26'] = f'{p2_in_progress/8*100:.0f}%'

# 保存文件
wb.save(file_path)
print(f'Excel 已更新：{file_path}')
print(f'Phase 1 完成：{p1_completed}/6 任务')
print(f'当前进行中：T07')
