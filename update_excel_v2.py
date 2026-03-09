from openpyxl import load_workbook

# 加载 Excel 文件
file_path = r'C:\Users\Administrator\openclaw-workspace\Small-Process-Calc\任务跟踪表.xlsx'
wb = load_workbook(file_path)
ws = wb['任务跟踪表']
ws_score = wb['积分统计']

# 重新计算积分（正确规则）
# T01-T06 的积分修正
updates = {
    2: 1.0,   # T01: 0.15/0.2 = 75% → +1.0
    3: 1.0,   # T02: 0.12/0.2 = 60% → +1.0
    4: 1.0,   # T03: 0.3/0.5 = 60% → +1.0
    5: 1.0,   # T04: 0.25/0.3 = 83% → +1.0
    6: 1.0,   # T05: 0.17/0.3 = 57% → +1.0
    7: 1.5,   # T06: 0.08/0.2 = 40% → +1.5 🚀
}

for row, score in updates.items():
    ws[f'I{row}'] = score  # 积分

# 计算当前总积分
total_score = sum(ws[f'I{row}'].value for row in range(2, 8) if ws[f'I{row}'].value)
ws_score['B16'] = total_score  # 当前总积分

# 更新完成率
completed = 6
ws_score['B19'] = f'{completed/35*100:.1f}%'

# 保存文件
wb.save(file_path)
print(f'Excel 已更新：{file_path}')
print(f'Phase 1 总积分：{total_score} 分（原 9.0 分 → 修正为 {total_score} 分）')
print(f'明细：T01-T05 各 +1.0 分，T06 +1.5 分（40% 提前完成）')
