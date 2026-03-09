from openpyxl import load_workbook

file_path = 'C:/Users/Administrator/openclaw-workspace/Small-Process-Calc/任务跟踪表.xlsx'
wb = load_workbook(file_path)
ws = wb['任务跟踪表']
ws_score = wb['积分统计']

import sys
sys.stdout.reconfigure(encoding='utf-8')

print('=== Task Tracker (Phase 1) ===')
for i in range(2, 8):
    task_id = ws[f'A{i}'].value
    task_name = ws[f'B{i}'].value
    status = ws[f'H{i}'].value
    score = ws[f'I{i}'].value
    print(f'{task_id} | {task_name} | {status} | Score: {score}')

print('\n=== Score Stats ===')
print(f'Total Score: {ws_score["B16"].value}')
print(f'Completion: {ws_score["B19"].value}')

print('\n=== Phase Progress ===')
for i in range(25, 32):
    phase = ws_score[f'A{i}'].value
    total = ws_score[f'B{i}'].value
    completed = ws_score[f'C{i}'].value
    progress = ws_score[f'D{i}'].value
    print(f'{phase} | Tasks: {total} | Done: {completed} | Progress: {progress}')
