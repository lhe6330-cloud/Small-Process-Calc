"""
Excel 报告导出模块
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def export_mode1_report(data: dict, input_params: dict) -> bytes:
    """导出模式 1 Excel 报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = '计算报告'
    
    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    title_font = Font(bold=True, size=14, color='00D4FF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 标题
    ws['A1'] = 'PDS CALC - 计算报告 (模式 1: 先加热再膨胀)'
    ws['A1'].font = title_font
    ws['A2'] = f'计算时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(size=10, color='94A3B8')
    
    # 输入参数
    ws['A4'] = '输入参数'
    ws['A4'].font = title_font
    
    headers = ['参数', '冷边', '热边', '涡轮']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # 数据
    rows = [
        ['介质', input_params['cold_side']['medium'], input_params['hot_side']['medium'], '-'],
        ['流量', f"{input_params['cold_side']['flow_rate']} {input_params['cold_side']['flow_unit']}", 
         f"{input_params['hot_side']['flow_rate']} {input_params['hot_side']['flow_unit']}", '-'],
        ['入口压力 (MPa.G)', input_params['cold_side']['p_in'], input_params['hot_side']['p_in'], '-'],
        ['出口压力 (MPa.G)', input_params['cold_side']['p_out'], input_params['hot_side']['p_out'], input_params['turbine']['p_out']],
        ['入口温度 (°C)', input_params['cold_side']['t_in'], input_params['hot_side']['t_in'], input_params['cold_side']['t_out']],
        ['出口温度 (°C)', input_params['cold_side']['t_out'], '-', '-'],
        ['绝热效率 (%)', '-', '-', input_params['turbine']['adiabatic_efficiency']],
    ]
    
    for row_idx, row in enumerate(rows, start=6):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 计算结果
    ws['A14'] = '计算结果'
    ws['A14'].font = title_font
    
    result_headers = ['项目', '数值', '单位']
    for col, h in enumerate(result_headers, start=1):
        cell = ws.cell(row=15, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    result_rows = [
        ['换热功率', f"{data['heat_exchanger']['q_power']:.2f}", 'kW'],
        ['热边出口温度', f"{data['heat_exchanger']['t_hot_out']:.1f}", '°C'],
        ['涡轮轴功率', f"{data['turbine']['power_shaft']:.2f}", 'kW'],
        ['发电功率', f"{data['turbine']['power_electric']:.2f}", 'kW'],
        ['涡轮出口温度', f"{data['turbine']['t_out']:.1f}", '°C'],
        ['电机选型', f"{data['selection']['motor']}", 'kW'],
        ['进口管道', f"DN{data['selection']['pipe_inlet']['recommended_dn']}", '-'],
        ['出口管道', f"DN{data['selection']['pipe_outlet']['recommended_dn']}", '-'],
        ['阀门', f"DN{data['selection']['valve']['valve_dn']}", '-'],
    ]
    
    for row_idx, row in enumerate(result_rows, start=16):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # 保存到字节
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def export_mode2_report(data: dict, input_params: dict) -> bytes:
    """导出模式 2 Excel 报告（先膨胀后回热）"""
    wb = Workbook()
    ws = wb.active
    ws.title = '计算报告 (模式 2)'
    
    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    title_font = Font(bold=True, size=14, color='00D4FF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 标题
    ws['A1'] = 'PDS CALC - 计算报告 (模式 2: 先膨胀后回热)'
    ws['A1'].font = title_font
    ws['A2'] = f'计算时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(size=10, color='94A3B8')
    
    # 输入参数
    ws['A4'] = '输入参数'
    ws['A4'].font = title_font
    
    turbine_in = input_params.get('turbine_in', {})
    turbine_params = input_params.get('turbine_params', {})
    hx_cold_out = input_params.get('hx_cold_out', {})
    hx_hot = input_params.get('hx_hot', {})
    
    headers = ['参数', '涡轮入口', '涡轮参数', '换热器冷边出口', '换热器热边']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    rows = [
        ['介质', turbine_in.get('medium', '-'), '-', '-', hx_hot.get('medium', '-')],
        ['流量', f"{turbine_in.get('flow_rate', 0)} {turbine_in.get('flow_unit', '')}", '-', '-', f"{hx_hot.get('flow_rate', 0)} {hx_hot.get('flow_unit', '')}"],
        ['入口压力 (MPa.G)', str(turbine_in.get('p_in', '-')), '-', str(hx_cold_out.get('p_out', '-')), str(hx_hot.get('p_in', '-'))],
        ['出口压力 (MPa.G)', str(turbine_params.get('p_out', '-')), '-', str(hx_cold_out.get('p_out', '-')), str(hx_hot.get('p_out', '-'))],
        ['入口温度 (°C)', str(turbine_in.get('t_in', '-')), '-', '-', str(hx_hot.get('t_in', '-'))],
        ['出口温度 (°C)', '-', '-', str(hx_cold_out.get('t_out', '-')), '-'],
        ['绝热效率 (%)', '-', str(turbine_params.get('adiabatic_efficiency', '-')), '-', '-'],
    ]
    
    for row_idx, row in enumerate(rows, start=6):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 计算结果
    ws['A14'] = '计算结果'
    ws['A14'].font = title_font
    
    result_headers = ['项目', '数值', '单位']
    for col, h in enumerate(result_headers, start=1):
        cell = ws.cell(row=15, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    result_rows = [
        ['涡轮轴功率', f"{data.get('turbine', {}).get('power_shaft', 0):.2f}", 'kW'],
        ['发电功率', f"{data.get('turbine', {}).get('power_electric', 0):.2f}", 'kW'],
        ['涡轮出口温度', f"{data.get('turbine', {}).get('t_out', 0):.1f}", '°C'],
        ['换热功率', f"{data.get('heat_exchanger', {}).get('q_power', 0):.2f}", 'kW'],
        ['热边出口温度', f"{data.get('heat_exchanger', {}).get('t_hot_out', 0):.1f}", '°C'],
        ['电机选型', f"{data.get('selection', {}).get('motor', 0)}", 'kW'],
        ['进口管道', f"DN{data.get('selection', {}).get('pipe_inlet', {}).get('recommended_dn', 0)}", '-'],
        ['出口管道', f"DN{data.get('selection', {}).get('pipe_outlet', {}).get('recommended_dn', 0)}", '-'],
        ['阀门', f"DN{data.get('selection', {}).get('valve', {}).get('valve_dn', 0)}", '-'],
    ]
    
    for row_idx, row in enumerate(result_rows, start=16):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # 保存到字节
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def export_mode3_report(data: dict, input_params: dict) -> bytes:
    """导出模式 3 Excel 报告（直接膨胀）"""
    wb = Workbook()
    ws = wb.active
    ws.title = '计算报告 (模式 3)'
    
    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    title_font = Font(bold=True, size=14, color='00D4FF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 标题
    ws['A1'] = 'PDS CALC - 计算报告 (模式 3: 直接膨胀)'
    ws['A1'].font = title_font
    ws['A2'] = f'计算时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(size=10, color='94A3B8')
    
    # 输入参数
    ws['A4'] = '输入参数'
    ws['A4'].font = title_font
    
    turbine_in = input_params.get('turbine_in', {})
    turbine_params = input_params.get('turbine_params', {})
    
    headers = ['参数', '数值']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    rows = [
        ['介质', turbine_in.get('medium', '-')],
        ['流量', f"{turbine_in.get('flow_rate', 0)} {turbine_in.get('flow_unit', '')}"],
        ['入口压力 (MPa.G)', str(turbine_in.get('p_in', '-'))],
        ['入口温度 (°C)', str(turbine_in.get('t_in', '-'))],
        ['出口压力 (MPa.G)', str(turbine_params.get('p_out', '-'))],
        ['绝热效率 (%)', str(turbine_params.get('adiabatic_efficiency', '-'))],
    ]
    
    for row_idx, row in enumerate(rows, start=6):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 计算结果
    ws['A13'] = '计算结果'
    ws['A13'].font = title_font
    
    result_headers = ['项目', '数值', '单位']
    for col, h in enumerate(result_headers, start=1):
        cell = ws.cell(row=14, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    result_rows = [
        ['涡轮轴功率', f"{data.get('turbine', {}).get('power_shaft', 0):.2f}", 'kW'],
        ['发电功率', f"{data.get('turbine', {}).get('power_electric', 0):.2f}", 'kW'],
        ['涡轮出口温度', f"{data.get('turbine', {}).get('t_out', 0):.1f}", '°C'],
        ['电机选型', f"{data.get('selection', {}).get('motor', 0)}", 'kW'],
        ['进口管道', f"DN{data.get('selection', {}).get('pipe_inlet', {}).get('recommended_dn', 0)}", '-'],
        ['出口管道', f"DN{data.get('selection', {}).get('pipe_outlet', {}).get('recommended_dn', 0)}", '-'],
        ['阀门', f"DN{data.get('selection', {}).get('valve', {}).get('valve_dn', 0)}", '-'],
    ]
    
    for row_idx, row in enumerate(result_rows, start=15):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    
    # 保存到字节
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
    """导出模式 1 Excel 报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = '计算报告'
    
    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    title_font = Font(bold=True, size=14, color='00D4FF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 标题
    ws['A1'] = '🍮 PDS CALC - 计算报告 (模式 1: 先加热再膨胀)'
    ws['A1'].font = title_font
    ws['A2'] = f'计算时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(size=10, color='94A3B8')
    
    # 输入参数
    ws['A4'] = '📝 输入参数'
    ws['A4'].font = title_font
    
    headers = ['参数', '冷边', '热边', '涡轮']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # 数据
    rows = [
        ['介质', input_params['cold_side']['medium'], input_params['hot_side']['medium'], '-'],
        ['流量', f"{input_params['cold_side']['flow_rate']} {input_params['cold_side']['flow_unit']}", 
         f"{input_params['hot_side']['flow_rate']} {input_params['hot_side']['flow_unit']}", '-'],
        ['入口压力 (MPa.G)', input_params['cold_side']['p_in'], input_params['hot_side']['p_in'], '-'],
        ['出口压力 (MPa.G)', input_params['cold_side']['p_out'], input_params['hot_side']['p_out'], input_params['turbine']['p_out']],
        ['入口温度 (°C)', input_params['cold_side']['t_in'], input_params['hot_side']['t_in'], input_params['cold_side']['t_out']],
        ['出口温度 (°C)', input_params['cold_side']['t_out'], '-', '-'],
        ['绝热效率 (%)', '-', '-', input_params['turbine']['adiabatic_efficiency']],
    ]
    
    for row_idx, row in enumerate(rows, start=6):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 计算结果
    ws['A14'] = '📊 计算结果'
    ws['A14'].font = title_font
    
    result_headers = ['项目', '数值', '单位']
    for col, h in enumerate(result_headers, start=1):
        cell = ws.cell(row=15, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    result_rows = [
        ['换热功率', f"{data['heat_exchanger']['q_power']:.2f}", 'kW'],
        ['热边出口温度', f"{data['heat_exchanger']['t_hot_out']:.1f}", '°C'],
        ['涡轮轴功率', f"{data['turbine']['power_shaft']:.2f}", 'kW'],
        ['发电功率', f"{data['turbine']['power_electric']:.2f}", 'kW'],
        ['涡轮出口温度', f"{data['turbine']['t_out']:.1f}", '°C'],
        ['电机选型', f"{data['selection']['motor']}", 'kW'],
        ['进口管道', f"DN{data['selection']['pipe_inlet']['recommended_dn']}", '-'],
        ['出口管道', f"DN{data['selection']['pipe_outlet']['recommended_dn']}", '-'],
        ['阀门', f"DN{data['selection']['valve']['valve_dn']}", '-'],
    ]
    
    for row_idx, row in enumerate(result_rows, start=16):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # 保存到字节
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
