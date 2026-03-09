"""
Excel Report Export Module - V2.0 (Mode 4 and Mode 5)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import io


def export_mode4_excel(data: dict, input_params: dict) -> bytes:
    """Export Mode 4 Excel Report (Separator Design)"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Separator Design'
    
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    title_font = Font(bold=True, size=14, color='00D4FF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = 'PDS CALC V2.0 - Separator Design Report'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    # Node Parameters
    ws['A4'] = 'Node Parameters'
    ws['A4'].font = title_font
    node_params = input_params.get('node_params', {})
    node_headers = ['Parameter', 'Value']
    for col, h in enumerate(node_headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    node_rows = [
        ['Position', input_params.get('node_id', '-')],
        ['Pressure (MPa.G)', node_params.get('p', '-')],
        ['Temperature (C)', node_params.get('t', '-')],
        ['Flow Rate', f"{node_params.get('flow_rate', 0)} {node_params.get('flow_unit', '')}"],
        ['Gas Density (kg/m3)', node_params.get('rho', '-')],
    ]
    for row_idx, row in enumerate(node_rows, start=6):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
    
    # VLE Result
    vle_result = data.get('vle', {})
    row_offset = 6
    if vle_result and not vle_result.get('skip', False):
        row_offset = len(node_rows) + 6
        ws[f'A{row_offset}'] = 'Vapor-Liquid Equilibrium'
        ws[f'A{row_offset}'].font = title_font
        vle_headers = ['Parameter', 'Value']
        for col, h in enumerate(vle_headers, start=1):
            cell = ws.cell(row=row_offset+1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        vle_rows = [
            ['Vapor Fraction', f"{vle_result.get('vapor_frac', 0) * 100:.1f} %"],
            ['Liquid Fraction', f"{vle_result.get('liquid_frac', 0) * 100:.1f} %"],
            ['Liquid Flow (T/h)', f"{vle_result.get('liquid_flow', 0):.2f}"],
        ]
        for row_idx, row in enumerate(vle_rows, start=row_offset+2):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
        row_offset += len(vle_rows) + 2
    
    # Separator Dimensions
    result_row = row_offset + 2
    ws[f'A{result_row}'] = 'Separator Dimensions'
    ws[f'A{result_row}'].font = title_font
    result_headers = ['Parameter', 'Value']
    for col, h in enumerate(result_headers, start=1):
        cell = ws.cell(row=result_row+1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    result_rows = [
        ['Diameter (mm)', data.get('diameter', 0)],
        ['Length (mm)', data.get('length', 0)],
        ['Residence Time (s)', f"{data.get('residence_time', 0):.1f}"],
        ['Check', 'OK' if data.get('check_passed') else 'WARN'],
    ]
    for row_idx, row in enumerate(result_rows, start=result_row+2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def export_mode5_excel(data: dict, input_params: dict) -> bytes:
    """Export Mode 5 Excel Report (Turbine 1D Design)"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Turbine Design'
    
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    title_font = Font(bold=True, size=14, color='00D4FF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = 'PDS CALC V2.0 - Turbine 1D Design Report'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    # Dimensions
    ws['A4'] = 'Dimensions'
    ws['A4'].font = title_font
    dim_headers = ['Parameter', 'Symbol', 'Value (mm)']
    for col, h in enumerate(dim_headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    dims = data.get('dimensions', {})
    dim_rows = [
        ['Impeller Outer Diameter', 'D1', dims.get('D1', 0)],
        ['Impeller Inner Diameter', 'D2', dims.get('D2', 0)],
        ['Inlet Blade Height', 'b1', dims.get('b1', 0)],
        ['Outlet Blade Height', 'b2', dims.get('b2', 0)],
        ['Blade Count', 'Z', dims.get('Z', 0)],
    ]
    for row_idx, row in enumerate(dim_rows, start=6):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
    
    # Velocity Triangle - Inlet
    ws['A12'] = 'Velocity Triangle - Inlet'
    ws['A12'].font = title_font
    vel_in_headers = ['Parameter', 'Value']
    for col, h in enumerate(vel_in_headers, start=1):
        cell = ws.cell(row=13, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    vel_in = data.get('velocity_triangle_in', {})
    vel_in_rows = [
        ['Absolute Velocity C1', f"{vel_in.get('C1', 0)} m/s"],
        ['Relative Velocity W1', f"{vel_in.get('W1', 0)} m/s"],
        ['Tangential Velocity U1', f"{vel_in.get('U1', 0)} m/s"],
        ['Absolute Flow Angle alpha1', f"{vel_in.get('alpha1', 0)} deg"],
        ['Relative Flow Angle beta1', f"{vel_in.get('beta1', 0)} deg"],
    ]
    for row_idx, row in enumerate(vel_in_rows, start=14):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
    
    # Velocity Triangle - Outlet
    ws['A20'] = 'Velocity Triangle - Outlet'
    ws['A20'].font = title_font
    vel_out_headers = ['Parameter', 'Value']
    for col, h in enumerate(vel_out_headers, start=1):
        cell = ws.cell(row=21, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    vel_out = data.get('velocity_triangle_out', {})
    vel_out_rows = [
        ['Absolute Velocity C2', f"{vel_out.get('C2', 0)} m/s"],
        ['Relative Velocity W2', f"{vel_out.get('W2', 0)} m/s"],
        ['Tangential Velocity U2', f"{vel_out.get('U2', 0)} m/s"],
        ['Absolute Flow Angle alpha2', f"{vel_out.get('alpha2', 0)} deg"],
        ['Relative Flow Angle beta2', f"{vel_out.get('beta2', 0)} deg"],
    ]
    for row_idx, row in enumerate(vel_out_rows, start=22):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
    
    # Performance
    ws['A28'] = 'Performance'
    ws['A28'].font = title_font
    perf_headers = ['Parameter', 'Value']
    for col, h in enumerate(perf_headers, start=1):
        cell = ws.cell(row=29, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    perf = data.get('performance', {})
    perf_rows = [
        ['Stage Efficiency', f"{data.get('thermo_params', {}).get('eta', 0)} %"],
        ['Calculated Power', f"{perf.get('P_calc', 0)} kW"],
        ['Input Power', f"{perf.get('P_input', 0)} kW"],
        ['Match', 'OK' if perf.get('match') else 'WARN'],
    ]
    for row_idx, row in enumerate(perf_rows, start=30):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
