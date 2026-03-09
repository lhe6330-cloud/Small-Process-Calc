import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建 Workbook
wb = Workbook()
ws = wb.active
ws.title = '任务跟踪表'

# 任务数据
tasks = [
    # Phase 1
    ['T01', '创建项目目录结构', 'Phase 1', 0.2, '', '', '', '⏳ 未开始', ''],
    ['T02', '初始化 Git 仓库', 'Phase 1', 0.2, '', '', '', '⏳ 未开始', ''],
    ['T03', '前端 Vue 3 项目搭建', 'Phase 1', 0.5, '', '', '', '⏳ 未开始', ''],
    ['T04', '后端 FastAPI 项目搭建', 'Phase 1', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T05', '工业仪表盘主题样式', 'Phase 1', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T06', '启动脚本编写', 'Phase 1', 0.2, '', '', '', '⏳ 未开始', ''],
    # Phase 2
    ['T07', '安装热力学库', 'Phase 2', 0.1, '', '', '', '⏳ 未开始', ''],
    ['T08', '压力单位转换工具', 'Phase 2', 0.2, '', '', '', '⏳ 未开始', ''],
    ['T09', 'H2O 物性计算模块', 'Phase 2', 0.5, '', '', '', '⏳ 未开始', ''],
    ['T10', '单一气体物性计算模块', 'Phase 2', 0.4, '', '', '', '⏳ 未开始', ''],
    ['T11', '混合介质物性计算模块', 'Phase 2', 0.5, '', '', '', '⏳ 未开始', ''],
    ['T12', '涡轮计算模块', 'Phase 2', 0.5, '', '', '', '⏳ 未开始', ''],
    ['T13', '换热计算模块', 'Phase 2', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T14', '模式 1 计算接口', 'Phase 2', 0.3, '', '', '', '⏳ 未开始', ''],
    # Phase 3
    ['T15', '加载标准数据 CSV', 'Phase 3', 0.2, '', '', '', '⏳ 未开始', ''],
    ['T16', '电机选型模块', 'Phase 3', 0.2, '', '', '', '⏳ 未开始', ''],
    ['T17', '管道通径计算模块', 'Phase 3', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T18', '蝶阀 Kv 数据搜集', 'Phase 3', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T19', '阀门 Kv 计算模块', 'Phase 3', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T20', '模式 2/3 计算接口', 'Phase 3', 0.2, '', '', '', '⏳ 未开始', ''],
    # Phase 4
    ['T21', '主布局组件', 'Phase 4', 0.4, '', '', '', '⏳ 未开始', ''],
    ['T22', '介质选择组件', 'Phase 4', 0.5, '', '', '', '⏳ 未开始', ''],
    ['T23', '流量/压力/温度输入组件', 'Phase 4', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T24', '模式 1 表单组件', 'Phase 4', 0.4, '', '', '', '⏳ 未开始', ''],
    ['T25', '模式 2/3 表单组件', 'Phase 4', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T26', '计算结果展示组件', 'Phase 4', 0.4, '', '', '', '⏳ 未开始', ''],
    ['T27', '表单校验逻辑', 'Phase 4', 0.2, '', '', '', '⏳ 未开始', ''],
    # Phase 5
    ['T28', 'Excel 报告导出', 'Phase 5', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T29', 'PDF 报告导出', 'Phase 5', 0.4, '', '', '', '⏳ 未开始', ''],
    ['T30', '报告 API 接口', 'Phase 5', 0.2, '', '', '', '⏳ 未开始', ''],
    # Phase 6
    ['T31', '前后端联调', 'Phase 6', 0.5, '', '', '', '⏳ 未开始', ''],
    ['T32', '边界条件测试', 'Phase 6', 0.4, '', '', '', '⏳ 未开始', ''],
    ['T33', '报告导出测试', 'Phase 6', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T34', '性能优化', 'Phase 6', 0.3, '', '', '', '⏳ 未开始', ''],
    ['T35', 'UI/UX 优化', 'Phase 6', 0.2, '', '', '', '⏳ 未开始', ''],
]

# 表头
headers = ['ID', '任务名称', '阶段', '计划工时 (h)', '开始时间', '完成时间', '实际工时 (h)', '状态', '积分']
ws.append(headers)

# 添加任务
for task in tasks:
    ws.append(task)

# 设置样式
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center')

# 表头样式
for col in range(1, 10):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# 列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 8

# 创建积分统计 Sheet
ws_score = wb.create_sheet(title='积分统计')

# 积分规则说明
ws_score['A1'] = '📊 积分规则'
ws_score['A1'].font = Font(bold=True, size=14, color='00D4FF')
ws_score['A3'] = '完成情况'
ws_score['B3'] = '积分'
ws_score['C3'] = '说明'

rules = [
    ['🚀 提前完成 (<50% 计划时间)', '+1.5', '实际工时 ≤ 0.5×计划工时'],
    ['✅ 按时完成 (≤计划时间)', '+1.0', '实际工时 ≤ 计划工时'],
    ['⚠️ 超时 0-100%', '+0.5', '计划工时 < 实际工时 ≤ 2×计划工时'],
    ['❌ 超时>100%', '-0.5', '实际工时 > 2×计划工时'],
]

for i, rule in enumerate(rules, start=4):
    ws_score[f'A{i}'] = rule[0]
    ws_score[f'B{i}'] = rule[1]
    ws_score[f'C{i}'] = rule[2]

# 当前积分统计
ws_score['A9'] = '📈 当前统计'
ws_score['A9'].font = Font(bold=True, size=14, color='00D4FF')
ws_score['A11'] = '统计项'
ws_score['B11'] = '数值'

stats = [
    ['总任务数', 35],
    ['已完成', 0],
    ['进行中', 0],
    ['未开始', 35],
    ['当前总积分', '=SUM(\'任务跟踪表\'!I2:I36)'],
    ['满分积分', 35],
    ['及格积分', 17.5],
    ['完成率', '0%'],
]

for i, stat in enumerate(stats, start=12):
    ws_score[f'A{i}'] = stat[0]
    ws_score[f'B{i}'] = stat[1]

# 阶段进度
ws_score['A22'] = '📋 阶段进度'
ws_score['A22'].font = Font(bold=True, size=14, color='00D4FF')
ws_score['A24'] = '阶段'
ws_score['B24'] = '任务数'
ws_score['C24'] = '已完成'
ws_score['D24'] = '进度'
ws_score['E24'] = '计划工时'
ws_score['F24'] = '实际工时'

phases = [
    ['Phase 1', 6, 0, '0%', 1.4, 0],
    ['Phase 2', 8, 0, '0%', 2.6, 0],
    ['Phase 3', 6, 0, '0%', 1.5, 0],
    ['Phase 4', 7, 0, '0%', 2.5, 0],
    ['Phase 5', 3, 0, '0%', 0.9, 0],
    ['Phase 6', 5, 0, '0%', 1.7, 0],
    ['总计', 35, 0, '0%', 7.0, 0],
]

for i, phase in enumerate(phases, start=25):
    ws_score[f'A{i}'] = phase[0]
    ws_score[f'B{i}'] = phase[1]
    ws_score[f'C{i}'] = phase[2]
    ws_score[f'D{i}'] = phase[3]
    ws_score[f'E{i}'] = phase[4]
    ws_score[f'F{i}'] = phase[5]

# 成就系统
ws_score['A33'] = '🏆 成就系统'
ws_score['A33'].font = Font(bold=True, size=14, color='00D4FF')
ws_score['A35'] = '成就'
ws_score['B35'] = '条件'
ws_score['C35'] = '状态'

achievements = [
    ['🚀 光速启动', 'Phase 1 全部按时完成', '🔒 锁定'],
    ['🔥 计算大师', 'Phase 2 积分≥6', '🔒 锁定'],
    ['⚡ 选型专家', 'Phase 3 积分≥4', '🔒 锁定'],
    ['🎨 界面达人', 'Phase 4 积分≥5', '🔒 锁定'],
    ['📄 报告高手', 'Phase 5 全部按时完成', '🔒 锁定'],
    ['🏅 完美收官', '总积分≥30', '🔒 锁定'],
]

for i, ach in enumerate(achievements, start=36):
    ws_score[f'A{i}'] = ach[0]
    ws_score[f'B{i}'] = ach[1]
    ws_score[f'C{i}'] = ach[2]

# 保存文件
file_path = r'C:\Users\Administrator\openclaw-workspace\Small-Process-Calc\任务跟踪表.xlsx'
wb.save(file_path)
print(f'Excel 文件已保存：{file_path}')
